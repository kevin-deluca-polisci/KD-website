#!/usr/bin/env python3
"""
Build the RA worksheet for candidacy events.

WHY A SEPARATE FILE FROM THE CSV. conditions/candidacy_events.csv is the
DESTINATION — the audited table the models read — and it is empty by design
until somebody has checked a row. The 128 drafted rows live in
drafts/candidacy_events_draft.csv, and opening the destination expecting to
find them is an easy and entirely reasonable mistake to make. So the coder gets
one workbook that already contains the draft, the columns to fill, and the
machine's evidence for its own guesses side by side.

Run:  python3 forecast/conditions/build_candidacy_worklist.py
Out:  forecast/conditions/candidacy_worklist.xlsx
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
DRAFT = HERE / "drafts" / "candidacy_events_draft.csv"
OUT = HERE / "candidacy_worklist.xlsx"

HEAD = Font(bold=True, color="FFFFFF")
HEADFILL = PatternFill("solid", fgColor="374151")
TOFILL = PatternFill("solid", fgColor="FEF3C7")     # amber: you fill this
CHECK = PatternFill("solid", fgColor="DBEAFE")      # blue: verify this
EVID = PatternFill("solid", fgColor="F3F4F6")       # grey: evidence, read only
WRAP = Alignment(wrap_text=True, vertical="top")

VERIFY = ["race_id", "person", "event_type", "known_by", "date_basis",
          "source_url"]
FILL = ["party", "is_incumbent", "event_date", "incumbent_on_ballot_after",
        "notes"]
DROP = ["bioguide"]          # never populated and nothing reads it

README = [
    ("Candidacy events — coding worksheet", ""),
    ("", ""),
    ("What this is",
     "One row per dated change in who is running for a House or Senate seat in "
     "2026. Two models on the site read it: Lockerbie counts open seats, "
     "Lewis-Beck & Quinlan counts Democratic Senate retirements. Both are "
     "currently frozen at one hand-typed number, which is why their historical "
     "lines are flat."),
    ("The point is WHEN, not just WHO",
     "A model run for 3 March 2026 may know about an announcement made on "
     "2 March and not one made on 4 March. That is the whole reason this file "
     "exists. Getting the dates casually right is worse than leaving them "
     "blank, because a wrong date looks exactly like a right one."),
    ("Where the rows came from",
     "Drafted automatically from Wikipedia revision history. EVERY DRAFTED "
     "VALUE IS A PROPOSAL, NOT AN ANSWER. The grey columns on the right are "
     "the machine's evidence for its own guess, so you can check it rather "
     "than trust it. They are ignored when the file is used."),
    ("", ""),
    ("BLUE columns", "Already filled by the scraper. Verify them against the "
                     "evidence columns and the source link. Fix what is wrong."),
    ("AMBER columns", "Blank. These are yours to code."),
    ("GREY columns", "Read-only evidence. Do not edit."),
    ("", ""),
    ("The field that matters most",
     "incumbent_on_ballot_after. FALSE if, after this event, the sitting "
     "member is NOT on the November ballot for this seat — the seat is open. "
     "TRUE if they are still running for it. This is what the models actually "
     "read, and it is deliberately separate from event_type because the "
     "mapping between them is not reliable. Code it from what is true. If it "
     "disagrees with event_type, the field wins and the disagreement goes in "
     "notes."),
    ("", ""),
    ("retiring", "Incumbent leaving public office -> FALSE"),
    ("seeking_other_office", "Incumbent running for something else -> FALSE. "
                             "Note: 'retiring to run for X' is this, NOT "
                             "'retiring'."),
    ("lost_primary", "Incumbent is gone even though the party still has a "
                     "nominee -> FALSE"),
    ("withdrew", "If a NON-incumbent withdrew, nothing about the incumbent "
                 "changed -> TRUE, or delete the row"),
    ("", ""),
    ("event_date",
     "The date the person ACTUALLY announced, from a news report or press "
     "release, as YYYY-MM-DD. Leave blank if you cannot find one — blank is "
     "fine and honest, because known_by still bounds it. Never guess. If the "
     "date is reported as a month with no day, leave it blank rather than "
     "writing the first: the first of the month looks like a real date to "
     "anything reading this file."),
    ("known_by", "DO NOT EDIT. The first date this appeared on Wikipedia — "
                 "the machine's honest upper bound on when it became public."),
    ("date_basis", "Leave as wiki_first_seen unless you filled event_date "
                   "from a real source, in which case change it to "
                   "'announcement'."),
    ("", ""),
    ("Announced then reversed",
     "Two separate rows, each with its own date. The models read in date "
     "order, so the reversal correctly re-closes the seat from its own date."),
    ("Resigned mid-term",
     "Code it, mark incumbent_on_ballot_after FALSE, and say so in notes — it "
     "matters for a special election."),
    ("Cannot tell if they are the incumbent",
     "Check the current member list. The _section column often says "
     "'Retirements' for a page region that also contains challengers."),
    ("", ""),
    ("When you are done",
     "Save the workbook. The blue and amber columns get exported to "
     "conditions/candidacy_events.csv — the audited table. Do not paste into "
     "that file by hand; it is generated."),
    ("How many", "128 drafted rows: 115 House, 6 Senate, spanning January 2025 "
                 "to August 2026."),
]


def main() -> int:
    if not DRAFT.exists():
        raise SystemExit(f"no draft at {DRAFT}")
    rows = list(csv.DictReader(DRAFT.open()))
    if not rows:
        raise SystemExit("draft is empty")

    evidence = [c for c in rows[0] if c.startswith("_")]
    order = VERIFY + FILL + evidence

    wb = Workbook()

    ws = wb.active
    ws.title = "Read me"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 104
    for i, (a, b) in enumerate(README, 1):
        ws.cell(i, 1, a).font = Font(bold=not b or a.isupper() or i == 1,
                                     size=14 if i == 1 else 11)
        c = ws.cell(i, 2, b)
        c.alignment = WRAP

    ws = wb.create_sheet("Candidacy events")
    for j, col in enumerate(order, 1):
        c = ws.cell(1, j, col)
        c.font = HEAD
        c.fill = HEADFILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = (
            34 if col in ("source_url", "notes") or col.startswith("_") else 19)
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, 2):
        for j, col in enumerate(order, 1):
            c = ws.cell(i, j, r.get(col, ""))
            c.alignment = WRAP
            if col in FILL:
                c.fill = TOFILL
            elif col in VERIFY:
                c.fill = CHECK
            else:
                c.fill = EVID

    # Dropdowns, so the two categorical fields cannot be typed three ways.
    n = len(rows) + 1
    for col, formula in (
            ("party", '"D,R,I"'),
            ("is_incumbent", '"TRUE,FALSE"'),
            ("incumbent_on_ballot_after", '"TRUE,FALSE"'),
            ("event_type", '"retiring,seeking_other_office,lost_primary,'
                           'withdrew,resigned"')):
        if col not in order:
            continue
        letter = get_column_letter(order.index(col) + 1)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{n}")

    wb.save(OUT)
    print(f"  wrote {OUT.relative_to(HERE.parents[1])}")
    print(f"  {len(rows)} rows | verify {len(VERIFY)} | code {len(FILL)} | "
          f"evidence {len(evidence)}")
    print(f"  dropped (never populated, nothing reads it): {DROP}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
